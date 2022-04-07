import openpyxl


class Excel:
    @staticmethod
    def writeExcel(lst):
        path = "\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
        wb = openpyxl.load_workbook(path)
        sh1 = wb['Sheet1']
        r = sh1.max_row
        c = sh1.max_column
        for j in range(1,c+1):
            sh1.cell(r+1,j).value = lst[j-1]

        path1="\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
        wb.save(path1)

    @staticmethod
    def editExcel():
        path = "\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
        wb = openpyxl.load_workbook(path)
        sh1 = wb['Sheet1']
        r = sh1.max_row
        c = sh1.max_column
        row = 1
        col = 1
        print("movie lists:")
        for i in range(2, r+1):
            val = sh1.cell(i, 1).value
            print(i, f".{val}")
        edit_movie = input("enter the movie to be edited:")

        # editing movie details
        for i in range(1,r+1):
            val = sh1.cell(i,1).value
            if edit_movie == val:
                for j in range(1,c+1):
                    edit_val = sh1.cell(row, col).value
                    sh1.cell(i, j, value=input(f"enter the {edit_val} to be edited: "))
                    col += 1
                break
        wb.save(path)

    @staticmethod
    def delExcel():
        path = "\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
        wb = openpyxl.load_workbook(path)
        sh1 = wb['Sheet1']
        r = sh1.max_row
        print("movie lists:")
        for i in range(2, r):
            val = sh1.cell(i, 1).value
            print(i, f".{val}")
        del_movie = input("enter the movie to be deleted:")

        # deleting movie details
        for i in range(1, r + 1):
            val = sh1.cell(i, 1).value
            if del_movie == val:
                sh1.delete_rows(i)
                break
        wb.save(path)




