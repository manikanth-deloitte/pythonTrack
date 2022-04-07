import openpyxl


class Register:
    def __init__(self, name, email, phone, age, password):
        self.name = name
        self.email = email
        self.phone = phone
        self.age = age
        self.password = password
        self.lst_details =[self.name, self.email, self.phone, self.age, self.password]

    def user_register(self):
        path = "\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
        wb = openpyxl.load_workbook(path)
        sh2 = wb['Sheet2']
        r = sh2.max_row
        c = sh2.max_column
        row = r+1
        for col in range(1,c+1):
            sh2.cell(row,col).value = self.lst_details[col-1]
        wb.save(path)

    def checkUser(self,name):
        path = "\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
        wb = openpyxl.load_workbook(path)
        sh2 = wb['Sheet2']
        r = sh2.max_row
        col=1
        for row in range(1, r + 1):
            user_name = sh2.cell(row, col).value
            if name == user_name:
                print("user registered already!")
                wb.save(path)
                break
        else:
            wb.save(path)
            self.user_register()


def userRegister():
    name = input("enter name: ")
    email = input("enter email-id: ")
    phone = input("enter phone number: ")
    age = input("enter age: ")
    password = input("enter password: ")
    reg = Register(name, email, phone, age, password)
    reg.checkUser(name)
