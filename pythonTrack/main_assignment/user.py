import openpyxl


class SeatsError(Exception):
    def __init__(self, data):
        self.data = data

    def __str__(self):
        return repr(self.data)


class User:
    def __init__(self):
        self.path = "\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
        self.wb = openpyxl.load_workbook(self.path)
        self.sh1 = self.wb['Sheet1']

    def todoAction(self):
        print("1.Book Tickets\n2.Cancel Tickets\n3.Give User Rating\n4.logout")

    def moviesList(self):
        r = self.sh1.max_row
        print("movie lists:")
        for i in range(2, r+1):
            val = self.sh1.cell(i, 1).value
            print(i, f".{val}")

    def movieDetails(self,opt):
        row = opt
        c = self.sh1.max_column
        mov_dic = {}
        for col in range(1,c+1):

            # add to dictionary
            attr = self.sh1.cell(1, col).value
            mov_dic[attr]=self.sh1.cell(row, col).value

            if col == 1 or col ==2 or col == 3 or col == 4 or col == 5 or col == 6:
                print(f"{attr}:",self.sh1.cell(row, col).value)
            if col == 12:
                print(f"{attr}:", self.sh1.cell(row, col).value)
            if col == c+1:
                print(f"{attr}:",self.sh1.cell(row, col).value)

        return mov_dic

    def updateCap(self,val,opt):
        row = opt
        c = self.sh1.max_column
        for col in range(1,c+1):
            if col == 13:
                self.sh1.cell(row, col).value =val
                self.wb.save(self.path)
                break

        # self.wb.save(self.path)
        print("up val:", self.sh1.cell(row, 13).value)

    def cancelTic(self,cancel_tic,opt):
        row = opt
        c = self.sh1.max_column
        seats = self.sh1.cell(row, 13).value
        update = int(seats)+cancel_tic
        for col in range(1, c + 1):
            if col == 13:
                self.sh1.cell(row, col).value = update
                self.wb.save(self.path)
                break
        # self.wb.save(self.path)
        print("up val:", self.sh1.cell(row, 13).value)

    def addUserRating(self,user_rating,opt):
        c = self.sh1.max_column
        self.sh1.cell(1, c+1).value = "User Rating"
        row = opt
        self.sh1.cell(row,c+1).value = user_rating
        self.wb.save(self.path)


def userAction(user1):
    user = User()
    print(f"******Welcome {user1} *******")
    user.moviesList()
    opt1 = int(input('enter movie: '))
    details = user.movieDetails(opt1)
    while True:
        user.todoAction()
        opt = int(input("select above options: "))
        if opt == 1:
            print(f"******Welcome {user1} *******")
            timings = details['Timings']
            lst = timings.split(',')
            dic_ti ={}
            print("movie timings : ")
            for i in range(len(lst)):
                print(i+1, ".", lst[i])
                dic_ti[i+1]=lst[i]

            # select timing
            sel = int(input("select timing: "))
            timing = dic_ti[sel]
            print("Timing:", timing)
            seats = details['Capacity']
            print("Remaining Seats:", seats)
            sel_seats = int(input("enter the no of seats to be selected:"))
            try:
                if sel_seats > int(seats):
                    raise SeatsError("selected seats should be less than seats available")
            except SeatsError as e:
                print(e.data)

            upt_seats = int(seats)-int(sel_seats)
            user.updateCap(upt_seats,opt1)
        elif opt == 2:
            print(f"******Welcome {user1} *******")
            cancel_tic = int(input("enter number of seats to be canceled: "))
            user.cancelTic(cancel_tic,opt1)
        elif opt == 3:
            print(f"******Welcome {user1} *******")
            user_rating = input(f"{user1} give rating for movie: ")
            user.addUserRating(user_rating,opt1)
        else:
            break


def userlogin(username,password):
    path = "\\Users\\amanikanth\\PycharmProjects\\pythonTrack\\main_assignment\\excelfiles\\movieDetails.xlsx"
    wb = openpyxl.load_workbook(path)
    sh2 = wb['Sheet2']
    r = sh2.max_row
    col = 1
    for row in range(1, r + 1):
        user_name = sh2.cell(row, col).value
        if username == user_name:
            pass_word = sh2.cell(row, 5).value

            if pass_word == password:
                print("logged in successful!")
                userAction(user_name)
                break
            else:
                print("password doesn't match")
                break
    else:
        print("register before logging in ")